#!/usr/bin/env python3
"""
电商利润表自动计算管线（模板）
====================
使用方式：
  1. 改 BASE 为项目数据目录路径
  2. 改 COST_PER_UNIT 为产品成本单价
  3. 运行: python3 pnl-pipeline.py
  4. 输出: ~/Desktop/hermes/<项目名>_利润表.xlsx

数据源要求：
  - 抖音：卓世数据/卓世销售数据/卓世抖音数据/月.csv（抖音小店结算CSV）
  - 京东：卓世数据/卓世销售数据/卓世京东数据/*.csv（营销对账表——仅优惠分摊）
  - 进销存：*进销售存出入库管理系统*.xlsx（出库明细表）
  - 拼多多：拼多多/拼多多刷单费用.xlsx

注意事项（见 SKILL.md Pitfalls 章节）：
  - 抖音退款字段为负数，需 abs()
  - 京东营销对账表不含实际售价，收入需用户补充
  - 拼多多费用表有"汇总"行需跳过
  - 进销存数量列在 col 9 (0-indexed)
"""

import csv, glob, os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ============ 配置 ============
BASE = '/Users/mac/Downloads/同步空间/抖音/直播项目/卓世雨刮器'
OUTPUT = '/Users/mac/Desktop/hermes/<项目名>_利润表.xlsx'
COST_PER_UNIT = 20  # 元/副（改为项目实际成本）


def parse_douyin_csv():
    """处理抖音CSV月销售数据"""
    dy_dir = f'{BASE}/卓世数据/卓世销售数据/卓世抖音数据'
    files = sorted(glob.glob(f'{dy_dir}/*.csv'))

    monthly = defaultdict(lambda: {
        'revenue': 0, 'platform_fee': 0, 'commission': 0,
        'refund': 0, 'net_revenue': 0, 'order_count': 0, 'refund_count': 0,
    })

    for f in files:
        fname = os.path.basename(f)
        parts = fname.replace('.csv', '').split('年')
        if len(parts) != 2:
            continue
        year = parts[0]
        month = parts[1].replace('月', '').strip()

        with open(f, 'r', encoding='utf-8-sig') as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                direction = row.get('动账方向', '')
                try:
                    fee = abs(float(row.get('平台服务费', '0') or 0))
                except:
                    fee = 0
                try:
                    comm = abs(float(row.get('佣金', '0') or 0))
                except:
                    comm = 0
                # 退款为负数，必须用 abs()
                try:
                    refund = abs(float(row.get('订单退款', '0') or 0))
                except:
                    refund = 0
                try:
                    pay = float(row.get('订单实付应结', '0') or 0)
                except:
                    pay = 0

                m = monthly[f'{year}_{month}']
                if direction == '入账':
                    m['revenue'] += pay
                    m['order_count'] += 1
                m['platform_fee'] += fee
                m['commission'] += comm
                if refund > 0:
                    m['refund'] += refund
                    m['refund_count'] += 1

    for m in monthly.values():
        m['net_revenue'] = m['revenue'] - m['refund']
    return monthly


def parse_jd_csv():
    """京东营销对账表——仅优惠分摊，不含实际售价"""
    return defaultdict(lambda: {'revenue': 0, 'fee': 0, 'net': 0})


def parse_inventory():
    """进销存出库总数"""
    files = glob.glob(f'{BASE}/*进销售存出入库管理系统*.xlsx')
    if not files:
        return 0
    wb = openpyxl.load_workbook(files[0], data_only=True)
    ws = wb['出库明细表']
    total = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        qty = row[9] if len(row) > 9 and isinstance(row[9], (int, float)) else 0
        total += qty
    return total


def parse_pdd_expenses():
    """拼多多费用（跳过汇总行）"""
    files = glob.glob(f'{BASE}/拼多多/*刷单费用*.xlsx') or glob.glob(f'{BASE}/拼多多/*费用*.xlsx')
    if not files:
        return {'总费用': 0, '直通车': 0, 'dsr_补单': 0, '评价_补单': 0,
                '保证金': 0, '快递助手': 0}
    wb = openpyxl.load_workbook(files[0], data_only=True)
    ws = wb[wb.sheetnames[0]]

    total = {'保证金': 0, '快递助手': 0, '直通车': 0,
             'dsr_补单': 0, '评价_补单': 0, '总费用': 0}

    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        # 跳过汇总行（日期非数字开头）
        date_val = str(row[0]).strip()
        if not date_val[:4].isdigit():
            continue

        fee_names = ['保证金', '快递助手', '充值小号', '活动保证金',
                     '直通车', '评价有礼', '信封费用']
        for name, val in zip(fee_names, [row[i] for i in range(1, 8)]):
            if val and isinstance(val, (int, float)):
                total[name] += val
                total['总费用'] += val

        dsr = row[11] if len(row) > 11 and isinstance(row[11], (int, float)) else 0
        pj = row[14] if len(row) > 14 and isinstance(row[14], (int, float)) else 0
        total['dsr_补单'] += dsr
        total['评价_补单'] += pj
        total['总费用'] += dsr + pj

    return total


def write_excel(dy_monthly, total_units, cogs, pdd):
    """输出利润表"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '利润总表'

    title_font = Font(name='微软雅黑', size=16, bold=True, color='1E2761')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E2761', end_color='1E2761', fill_type='solid')
    data_font = Font(name='微软雅黑', size=10)
    money_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7'))

    ws.merge_cells('A1:G1')
    ws['A1'] = '利润表'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    rev_total = sum(m['net_revenue'] for m in dy_monthly.values())
    fee_total = sum(m['platform_fee'] for m in dy_monthly.values())
    comm_total = sum(m['commission'] for m in dy_monthly.values())
    gp = rev_total - cogs
    np_val = gp - fee_total - comm_total - pdd['总费用']

    headers = ['项目', '金额（元）', '说明']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    items = [
        ('营业收入', rev_total, '抖音小店结算净收入'),
        ('减：营业成本', -cogs, f'出库{total_units}件 × {COST_PER_UNIT}元/件'),
        ('= 毛利', gp, f'毛利率 {gp/rev_total*100:.1f}%' if rev_total else ''),
        ('减：平台服务费', -fee_total, '抖音平台扣点'),
        ('减：佣金', -comm_total, '达人佣金'),
        ('减：拼多多费用', -pdd['总费用'],
         f'直通车{pdd["直通车"]:.0f}/DSR补单{pdd["dsr_补单"]:.0f}/评价补单{pdd["评价_补单"]:.0f}'),
        ('= 净利润', np_val, f'净利率 {np_val/rev_total*100:.1f}%' if rev_total else ''),
    ]

    for r_offset, (name, val, note) in enumerate(items):
        r = 5 + r_offset
        ws.cell(row=r, column=1, value=name).font = data_font
        ws.cell(row=r, column=1).border = thin_border
        if val is not None:
            ws.cell(row=r, column=2, value=val).font = data_font
            ws.cell(row=r, column=2).number_format = money_fmt
        ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=3, value=note).font = Font(name='微软雅黑', size=9, color='7F8C8D')
        ws.cell(row=r, column=3).border = thin_border

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 50

    # Sheet 2: 月度明细
    ws2 = wb.create_sheet('抖音月度明细')
    ws2.merge_cells('A1:H1')
    ws2['A1'] = '抖音月度收入明细'
    ws2['A1'].font = title_font
    headers2 = ['月份', '订单数', '收入(实付)', '退款', '平台服务费', '佣金', '净收入', '退款率']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, (ym, m) in enumerate(sorted(dy_monthly.items())):
        r = 4 + i
        ws2.cell(row=r, column=1, value=ym).font = data_font
        ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=2, value=m['order_count']).font = data_font
        for col, key in [(3, 'revenue'), (4, 'refund'), (5, 'platform_fee'),
                         (6, 'commission'), (7, 'net_revenue')]:
            cell = ws2.cell(row=r, column=col, value=round(m[key], 2))
            cell.font = data_font
            cell.number_format = money_fmt
        ref_rate = m['refund'] / m['revenue'] if m['revenue'] else 0
        ws2.cell(row=r, column=8, value=round(ref_rate, 4))
        ws2.cell(row=r, column=8).number_format = pct_fmt
        for col in range(1, 9):
            ws2.cell(row=r, column=col).border = thin_border

    wb.save(OUTPUT)
    print(f'✅ 已保存: {OUTPUT}')


if __name__ == '__main__':
    print('=== 电商利润表自动计算 ===')
    dy = parse_douyin_csv()
    rev = sum(m['net_revenue'] for m in dy.values())
    print(f'抖音收入: {rev:>.2f}（{len(dy)}个月）')

    units = parse_inventory()
    cogs = units * COST_PER_UNIT
    print(f'产品成本: {cogs:.2f}（{units}件 × {COST_PER_UNIT}元）')

    pdd = parse_pdd_expenses()
    print(f'拼多多费用: {pdd["总费用"]:.2f}')

    fee = sum(m['platform_fee'] for m in dy.values())
    comm = sum(m['commission'] for m in dy.values())
    gp = rev - cogs
    np = gp - fee - comm - pdd['总费用']
    print(f'净利润: {np:.2f}（毛利率 {gp/rev*100:.1f}%，净利率 {np/rev*100:.1f}%）')

    write_excel(dy, units, cogs, pdd)
